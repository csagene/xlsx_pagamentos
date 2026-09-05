import streamlit as st
import pandas as pd
import numpy as np
import io
import json
import os

# ---------------------------------------------------------
# CONFIGURAÇÃO GERAL
# ---------------------------------------------------------
st.set_page_config(page_title="Sistema de Relatórios INAS", page_icon="📊", layout="wide")

st.markdown('''
    <style>
    header[data-testid="stHeader"] { background: transparent !important; }
    section[data-testid="stSidebar"] { background-color: #0d0d0d !important; }
    section[data-testid="stSidebar"] .stMarkdown p,
    section[data-testid="stSidebar"] .stMarkdown h1,
    section[data-testid="stSidebar"] .stMarkdown h2,
    section[data-testid="stSidebar"] .stMarkdown h3,
    section[data-testid="stSidebar"] div[data-testid="stMarkdownContainer"] p,
    section[data-testid="stSidebar"] .stRadio label,
    section[data-testid="stSidebar"] div[role="radiogroup"] div { color: #f0f0f0 !important; }
    .block-container { padding-bottom: 100px !important; padding-top: 1rem !important; }
    div[data-testid="stHorizontalBlock"]:has(~ div #nav-buttons-hook) {
        position: fixed; bottom: 0; left: 0; width: 100%;
        background-color: var(--background-color); padding: 15px 30px; z-index: 999;
        border-top: 1px solid var(--secondary-background-color); box-shadow: 0 -4px 6px -1px rgba(0,0,0,0.05); margin: 0;
    }
    @media (min-width: 50.625rem) {
        div[data-testid="stHorizontalBlock"]:has(~ div #nav-buttons-hook) { padding-left: 21rem; padding-right: 2rem; }
    }
    
    /* Make the filters expander sticky at the top */
    div[data-testid="stExpander"]:has(#filters-sticky-hook) {
        position: sticky;
        top: 0.5rem;
        z-index: 99999;
        background-color: var(--secondary-background-color);
        box-shadow: 0 4px 6px -1px rgba(0,0,0,0.05);
        border-radius: 0.5rem;
        overflow: visible !important;
    }
    div[data-testid="stExpander"]:has(#filters-sticky-hook) > div {
        overflow: visible !important;
    }
    div[data-testid="stExpanderDetails"] {
        overflow: visible !important;
    }

    /* ===== SPINNER NATIVO BLOQUEANTE (FULL-PAGE) ===== */
    /* Quando o Streamlit activa o st.spinner, mostramos a todo o ecrã */
    div[data-testid="stSpinner"] {
        position: fixed !important;
        top: 0 !important;
        left: 0 !important;
        width: 100vw !important;
        height: 100vh !important;
        background: rgba(0, 0, 0, 0.65) !important;
        backdrop-filter: blur(4px) !important;
        z-index: 999999 !important;
        display: flex !important;
        align-items: center !important;
        justify-content: center !important;
    }
    
    div[data-testid="stSpinner"] > div {
        background: #1e1e2e !important;
        padding: 40px 60px !important;
        border-radius: 16px !important;
        box-shadow: 0 25px 60px rgba(0,0,0,0.5) !important;
        color: white !important;
        display: flex !important;
        flex-direction: column !important;
        align-items: center !important;
        gap: 15px !important;
    }
    
    /* Aumentar um pouco o círculo animado */
    div[data-testid="stSpinner"] > div > div:first-child {
        transform: scale(1.5) !important;
        margin-bottom: 10px !important;
    }
    
    div[data-testid="stSpinner"] p {
        color: #e0e0ff !important;
        font-size: 1.1rem !important;
        font-weight: 500 !important;
        letter-spacing: 0.03em !important;
        margin: 0 !important;
    }
    /* ====================================== */
    </style>
''', unsafe_allow_html=True)


TEMPLATES_FILE = "templates.json"

DEFAULT_TEMPLATES = {
    "modelo de tabela globalizadovf_final_xls": {
        "colunas_agrupamento": ["Ano ", "Mês", "Província", "Delegação", "Distrito", "Fonte", "Programa", "Implementador", "Provedor  servico"],
        "colunas_metricas": ["F", "M", "Benef. Distintos", "1x", "2x", "3x", "4+", "Pagamentos", "Valor Pago"],
        "mensal_rename_map": {
            "Ano ": "Ano", "Mês": "Mes", "Provedor  servico": "Provedor servico"
        }
    }
}

def carregar_templates():
    if not os.path.exists(TEMPLATES_FILE):
        with open(TEMPLATES_FILE, "w", encoding="utf-8") as f:
            json.dump(DEFAULT_TEMPLATES, f, ensure_ascii=False, indent=4)
        return DEFAULT_TEMPLATES
    
    with open(TEMPLATES_FILE, "r", encoding="utf-8") as f:
        try:
            return json.load(f)
        except:
            return DEFAULT_TEMPLATES

def guardar_templates(templates):
    with open(TEMPLATES_FILE, "w", encoding="utf-8") as f:
        json.dump(templates, f, ensure_ascii=False, indent=4)

if "templates" not in st.session_state:
    st.session_state.templates = carregar_templates()

# Inicializar variáveis na memória
if "df" not in st.session_state:
    st.session_state.df = None
if "df_editado" not in st.session_state:
    st.session_state.df_editado = None
if "relatorio_final" not in st.session_state:
    st.session_state.relatorio_final = None
if "col_agrupamento" not in st.session_state:
    st.session_state.col_agrupamento = []

# ---------------------------------------------------------
# MENU DE NAVEGAÇÃO LATERAL
# ---------------------------------------------------------
PAGINAS = [
    "📥 1. Carregar & Editar Dados", 
    "📑 2. Gerar Relatório", 
    "📈 3. Dashboard Visual"
]

if "pagina_atual" not in st.session_state:
    st.session_state.pagina_atual = PAGINAS[0]

def ir_proximo():
    if st.session_state.pagina_atual in PAGINAS:
        idx = PAGINAS.index(st.session_state.pagina_atual)
        if idx < len(PAGINAS) - 1:
            st.session_state.pagina_atual = PAGINAS[idx + 1]

def ir_anterior():
    if st.session_state.pagina_atual in PAGINAS:
        idx = PAGINAS.index(st.session_state.pagina_atual)
        if idx > 0:
            st.session_state.pagina_atual = PAGINAS[idx - 1]

st.sidebar.markdown("### 🧭 Organização")
st.sidebar.markdown("Navegue entre as páginas da aplicação:")

idx = PAGINAS.index(st.session_state.pagina_atual) if st.session_state.pagina_atual in PAGINAS else 0
pagina = st.sidebar.radio("Menu de Páginas", PAGINAS, index=idx, label_visibility="collapsed")

if pagina != st.session_state.pagina_atual:
    st.session_state.pagina_atual = pagina
    st.rerun()

st.sidebar.divider()
if st.session_state.df_editado is not None:
    st.sidebar.success(f"✅ Dados carregados ({len(st.session_state.df_editado)} registos)")
else:
    st.sidebar.warning("⚠️ Nenhum dado carregado")
    
if st.sidebar.button("🔄 Limpar Cache / Recarregar App", width="stretch"):
    for key in list(st.session_state.keys()):
        del st.session_state[key]
    st.rerun()

def adicionar_linha_totais(df_resultado, colunas_agrupamento, is_cumulativo=False, df_bruto=None, meta=None):
    df_resultado = df_resultado.copy()
    primeira_col = colunas_agrupamento[0] if colunas_agrupamento and colunas_agrupamento[0] in df_resultado.columns else df_resultado.columns[0]
        
    def extract_unique_items(series):
        unique_items = set()
        for val in series.dropna().astype(str):
            for item in val.split(','):
                item = item.strip()
                if item and item != "N/D" and item.lower() != "fonte_financiamento":
                    unique_items.add(item)
        lista = list(unique_items)
        try:
            lista.sort(key=float)
        except ValueError:
            lista.sort()
        return ", ".join(lista)
        
    col_fonte = None
    col_impl = None
    col_mes = None
    
    for col in df_resultado.columns:
        c_lower = str(col).lower()
        if "fonte" in c_lower:
            col_fonte = col
        if "implementador" in c_lower:
            col_impl = col
        if "mes" in c_lower or "mês" in c_lower:
            col_mes = col
            
    val_impl = "PMA" if st.session_state.get("modelo_selecionado") == "PMA" else ("" if st.session_state.get("modelo_selecionado") == "INAS" else "INAS")
    if col_impl:
        df_resultado[col_impl] = val_impl
        
    linhas_totais = []
    
    t_geral = {}
    has_set_geral = False
    for col in df_resultado.columns:
        if col == col_impl:
            t_geral[col] = val_impl
        elif any(normalize_text(col) == normalize_text(c) for c in colunas_agrupamento):
            if not has_set_geral:
                t_geral[col] = "TOTAL"
                has_set_geral = True
            else:
                unique_vals = [v for v in df_resultado[col].dropna().astype(str).unique() if v.strip() != ""]
                t_geral[col] = ", ".join(unique_vals)
        elif pd.api.types.is_numeric_dtype(df_resultado[col]) and not any(normalize_text(col) == normalize_text(c) for c in colunas_agrupamento):
            is_distinct = False
            c_lower = str(col).strip().lower()
            val = str(col).strip().upper()
            
            # Check if this column is a distinct counting metric
            if "benef" in c_lower or val.startswith('F') or val.startswith('M') or val.endswith('X') or val.endswith('X_ACUM') or val.startswith('F_ACUM') or val.startswith('M_ACUM'):
                is_distinct = True
                
            if is_distinct and df_bruto is not None and meta is not None:
                col_b = meta['col_beneficiario']
                col_s = meta['col_sexo']
                if "benef" in c_lower:
                    t_geral[col] = df_bruto[col_b].nunique()
                elif (val.startswith('F') or val.startswith('F_ACUM')) and col_s:
                    t_geral[col] = df_bruto[df_bruto[col_s].astype(str).str.upper().str.startswith('F')][col_b].nunique()
                elif (val.startswith('M') or val.startswith('M_ACUM')) and col_s:
                    t_geral[col] = df_bruto[df_bruto[col_s].astype(str).str.upper().str.startswith('M')][col_b].nunique()
                elif val.endswith('X') or val.endswith('X_ACUM'):
                    v_clean = val.replace("_ACUM", "")
                    target_freq = int(v_clean.replace("X", ""))
                    if 'vezes' in df_bruto.columns:
                        max_freqs = df_bruto.groupby(col_b)['vezes'].max()
                        t_geral[col] = (max_freqs == target_freq).sum()
                    else:
                        t_geral[col] = df_bruto[df_bruto['categoria_vezes'] == v_clean][col_b].nunique()
            else:
                if is_cumulativo:
                    if col_mes:
                        agrup_sem_mes = [c for c in colunas_agrupamento if c != col_mes and c in df_resultado.columns]
                        if agrup_sem_mes:
                            t_geral[col] = df_resultado.groupby(agrup_sem_mes)[col].max().sum()
                        else:
                            t_geral[col] = df_resultado[col].max()
                    else:
                        t_geral[col] = df_resultado[col].max()
                else:
                    t_geral[col] = df_resultado[col].sum()
        else:
            if col not in t_geral:
                unique_vals = [v for v in df_resultado[col].dropna().astype(str).unique() if v.strip() != ""]
                t_geral[col] = ", ".join(unique_vals)
            
    linhas_totais.append(t_geral)
    
    df_totais = pd.DataFrame(linhas_totais)
    df_completo = pd.concat([df_resultado, df_totais], ignore_index=True)
    return df_completo, df_totais
import unicodedata

def normalize_text(text):
    if not isinstance(text, str):
        return ""
    text = unicodedata.normalize('NFKD', text).encode('ASCII', 'ignore').decode('utf-8')
    return text.lower().strip()

def processar_relatorio(df, template):
    colunas_df = df.columns.tolist()
    col_agrupamento = template["colunas_agrupamento"]
    col_metricas = template["colunas_metricas"]
    
    if not col_agrupamento:
        raise ValueError("O template deve ter pelo menos uma coluna de agrupamento.")
        
    aliases = {
        "ano": ["ano_pagamento", "ano", "year"],
        "mês": ["meses_pagamento", "mes", "mês", "meses", "month"],
        "província": ["provincia", "província", "province"],
        "delegação": ["delegacao", "delegação"],
        "distrito": ["distrito", "district"],
        "fonte": ["fonte_financiamento", "fonte", "source"],
        "programa": ["programa_social", "programa"],
        "implementador": ["implementador", "ps_name", "implementer"],  
        "provedor": ["psp_name", "provedor", "operador"]
    }
        
    col_agrupamento_reais = []
    rename_dict_seguro = {}
    for col_temp in col_agrupamento:
        temp_norm = normalize_text(col_temp)
        col_encontrada = None
        
        # 1. Exact match (normalized)
        for col in colunas_df:
            if normalize_text(col) == temp_norm:
                col_encontrada = col
                break
                
        # 2. Alias match
        if not col_encontrada:
            for key, alias_list in aliases.items():
                if key in temp_norm:
                    for alias in alias_list:
                        for col in colunas_df:
                            if alias == normalize_text(col):
                                col_encontrada = col
                                break
                        if col_encontrada: break
                if col_encontrada: break
                
        # 3. Substring match
        if not col_encontrada:
            for col in colunas_df:
                col_norm = normalize_text(col)
                if len(temp_norm) >= 3 and temp_norm in col_norm:
                    col_encontrada = col
                    break

        if col_encontrada:
            col_agrupamento_reais.append(col_encontrada)
            rename_dict_seguro[col_encontrada] = col_temp
        else:
            df[col_temp] = "N/D"
            col_agrupamento_reais.append(col_temp)
            
    def detetar_coluna(lista_colunas, palavras_chave):
        for col in lista_colunas:
            if str(col).lower().strip() in palavras_chave: return col
        for col in lista_colunas:
            for p in palavras_chave:
                if p in str(col).lower(): return col
        return None
        
    if 'Beneficiario' in colunas_df:
        col_beneficiario = 'Beneficiario'
    else:
        col_beneficiario = detetar_coluna(colunas_df, ['código', 'codigo', 'beneficiario', 'beneficiário', 'bi', 'nuit', 'agregado'])
        
    if 'Sexo' in colunas_df:
        col_sexo = 'Sexo'
    else:
        col_sexo = detetar_coluna(colunas_df, ['sexo', 'genero', 'género', 'sex'])
        
    if 'Valores_Pagos' in colunas_df:
        col_valor = 'Valores_Pagos'
    elif 'Valor Pago' in colunas_df:
        col_valor = 'Valor Pago'
    else:
        col_valor = detetar_coluna(colunas_df, ['valores_pagos', 'valores pagos', 'valor pago', 'valor_pago', 'valor', 'pago', 'montante', 'dinheiro', 'total', 'unnamed'])
    
    if not col_beneficiario: col_beneficiario = colunas_df[0]
    if not col_valor: col_valor = colunas_df[-1]
    
    def clean_currency(val):
        if pd.isna(val): return 0.0
        if isinstance(val, (int, float)): return float(val)
        s = str(val).upper().replace("MT", "").replace("MZN", "").strip()
        s = s.replace(" ", "")
        if '.' in s and ',' in s:
            if s.rfind(',') > s.rfind('.'):
                s = s.replace('.', '').replace(',', '.')
            else:
                s = s.replace(',', '')
        elif ',' in s:
            s = s.replace(',', '.')
        try:
            return float(s)
        except:
            return 0.0
            
    if col_valor in df.columns:
        df[col_valor] = df[col_valor].apply(clean_currency)
        
    # Preencher campos (Delegação, Distrito, etc) vazios em função do número de telefone / beneficiário
    import numpy as np
    col_telefone = detetar_coluna(colunas_df, ['telefone', 'numro', 'numero', 'conta', 'celular', 'contacto', 'phone'])
    if not col_telefone:
        col_telefone = col_beneficiario
        
    if col_telefone and col_telefone in df.columns:
        cols_to_fill = []
        keys_to_fill = ["delegação", "delegacao", "distrito", "provedor"]
        
        for key in keys_to_fill:
            for r_col in colunas_df:
                if normalize_text(r_col) in aliases.get(key, [key]):
                    if r_col not in cols_to_fill: cols_to_fill.append(r_col)
                    
        for r_col, f_col in template.get("mapeamento", {}).items():
            for key in keys_to_fill:
                if key in f_col.lower() and r_col in colunas_df and r_col not in cols_to_fill:
                    cols_to_fill.append(r_col)
                    
        if cols_to_fill:
            for c in cols_to_fill:
                df[c] = df[c].replace(r'^\s*$', np.nan, regex=True)
            
            # Apenas aplicar a transformacao onde o telefone não é nulo para evitar erro no groupby
            mask_valida = df[col_telefone].notna() & (df[col_telefone].astype(str).str.strip() != "")
            if mask_valida.any():
                df.loc[mask_valida, cols_to_fill] = df.loc[mask_valida].groupby(col_telefone)[cols_to_fill].transform(lambda x: x.ffill().bfill())
                
            for c in cols_to_fill:
                df[c] = df[c].fillna("")
        
    # Agrupamento básico
    def join_unique(series):
        return ", ".join([str(x) for x in series.dropna().unique() if str(x).strip() != ""])
        
    # Detectar se já existe uma coluna de "Número de Pagamentos" (ex: N_Pagamentos)
    col_pagamentos_raw = None
    for col in colunas_df:
        if normalize_text(col) in ["n_pagamentos", "n pagamentos", "npagamentos", "pagamentos"]:
            col_pagamentos_raw = col
            break
            
    if col_pagamentos_raw:
        df[col_pagamentos_raw] = pd.to_numeric(df[col_pagamentos_raw], errors='coerce').fillna(1).astype(int)
        agg_pag = pd.NamedAgg(column=col_pagamentos_raw, aggfunc='sum')
    else:
        agg_pag = pd.NamedAgg(column=col_beneficiario, aggfunc='count')
        
    agg_dict = {
        'PAGAMENTOS': agg_pag,
        'VALOR_PAGO': pd.NamedAgg(column=col_valor, aggfunc='sum'),
        'BENEF_DISTINTOS': pd.NamedAgg(column=col_beneficiario, aggfunc='nunique')
    }
    
    colunas_string_join = template.get("colunas_string_join", [])
    for c_final in colunas_string_join:
        encontrou = False
        for r_col, f_col in template.get("mapeamento", {}).items():
            if f_col == c_final and r_col in df.columns:
                agg_dict[c_final] = pd.NamedAgg(column=r_col, aggfunc=join_unique)
                encontrou = True
                break
        if not encontrou:
            # If not in mapeamento, try to find a matching column directly
            c_norm = normalize_text(c_final)
            col_encontrada = None
            for col in colunas_df:
                if normalize_text(col) == c_norm:
                    col_encontrada = col
                    break
            if not col_encontrada:
                for key, alias_list in aliases.items():
                    if key in c_norm:
                        for alias in alias_list:
                            for col in colunas_df:
                                if alias == normalize_text(col):
                                    col_encontrada = col
                                    break
                            if col_encontrada: break
                    if col_encontrada: break
            if col_encontrada:
                agg_dict[c_final] = pd.NamedAgg(column=col_encontrada, aggfunc=join_unique)
                
    resumo_basico = df.groupby(col_agrupamento_reais).agg(**agg_dict)
    
    resumo_basico['Pagamentos'] = resumo_basico['PAGAMENTOS']
    resumo_basico['Valor Pago'] = resumo_basico['VALOR_PAGO']
    resumo_basico['Benef. Distintos'] = resumo_basico['BENEF_DISTINTOS']
    
    if col_sexo:
        df_sexo_base = df.copy()
        df_sexo_base[col_sexo] = df_sexo_base[col_sexo].fillna("SEM_GENERO").replace("", "SEM_GENERO")
        agrupamento_sexo = list(dict.fromkeys(col_agrupamento_reais + [col_beneficiario, col_sexo]))
        df_sexo = df_sexo_base.drop_duplicates(subset=agrupamento_sexo)
        sexo_pivot = pd.pivot_table(df_sexo, index=col_agrupamento_reais, columns=col_sexo, values=col_beneficiario, aggfunc='nunique', fill_value=0)
        
        for sexo_col in sexo_pivot.columns:
            val = str(sexo_col).strip().upper()
            if val.startswith('F'):
                resumo_basico['F'] = sexo_pivot[sexo_col]
            elif val.startswith('M'):
                resumo_basico['M'] = sexo_pivot[sexo_col]
            else:
                nome_col = f"Sexo_{val}" if val != "SEM_GENERO" else "Sem_Gênero"
                resumo_basico[nome_col] = sexo_pivot[sexo_col]
                
    # Abordagem 1: Frequência Progressiva (Parcelas)
    df_freq = df.copy()
    colunas_ordenacao = [c for c in col_agrupamento_reais if any(k in normalize_text(c) for k in ["ano", "mes", "mês", "data"])]
    if colunas_ordenacao:
        df_freq = df_freq.sort_values(by=colunas_ordenacao)
        
    df_freq['vezes'] = df_freq.groupby(col_beneficiario).cumcount() + 1
    df_freq['categoria_vezes'] = df_freq['vezes'].astype(str) + 'X'
    freq_pivot = pd.pivot_table(df_freq, index=col_agrupamento_reais, columns='categoria_vezes', values=col_beneficiario, aggfunc='nunique', fill_value=0)
    
    relatorio_final = resumo_basico.join(freq_pivot).fillna(0).reset_index()
    
    rename_dict = rename_dict_seguro
    relatorio_final = relatorio_final.rename(columns=rename_dict)
    
    for metrica in col_metricas:
        if metrica not in relatorio_final.columns:
            relatorio_final[metrica] = 0
            
    colunas_string_join = template.get("colunas_string_join", [])
    for c in colunas_string_join:
        if c not in relatorio_final.columns:
            relatorio_final[c] = ""
    ordem_final = col_agrupamento + colunas_string_join + [m for m in col_metricas if m in relatorio_final.columns]
    relatorio_final = relatorio_final[ordem_final]
    
    for col in relatorio_final.columns:
        if "ano" in str(col).lower():
            # Converter para string e remover '.0' de floats para não aparecer com vírgulas/decimais na UI
            relatorio_final[col] = relatorio_final[col].astype(str).str.replace(r'\.0$', '', regex=True).replace('nan', '')
            
    df_mensal = relatorio_final.copy()
    df_cumulativo = relatorio_final.copy()
    
    col_mes_real = None
    col_mes_final = None
    for r_col, f_col in zip(col_agrupamento_reais, col_agrupamento):
        if 'mes' in f_col.lower() or 'mês' in f_col.lower():
            col_mes_real = r_col
            col_mes_final = f_col
            break
            
    if col_mes_real:
        def extract_num(val):
            try:
                import re
                val_str = str(val).lower()
                nums = re.findall(r'\d+', val_str)
                if nums: return float(nums[0])
                meses_map = {'jan': 1, 'fev': 2, 'mar': 3, 'abr': 4, 'mai': 5, 'jun': 6, 'jul': 7, 'ago': 8, 'set': 9, 'out': 10, 'nov': 11, 'dez': 12}
                for mes_chave, num in meses_map.items():
                    if mes_chave in val_str: return float(num)
                return 0.0
            except:
                return 0.0
                
        df_sort = df.copy()
        df_sort['_sort_mes'] = df_sort[col_mes_real].apply(extract_num)
        
        col_agrup_sem_mes_reais = [c for c in col_agrupamento_reais if c != col_mes_real]
        col_agrup_sem_mes_finais = [c for c in col_agrupamento if c != col_mes_final]
        
        df_sort = df_sort.sort_values(by=col_agrup_sem_mes_reais + ['_sort_mes'])
        
        lista_cumul = []
        for chaves, grupo in df_sort.groupby(col_agrup_sem_mes_reais):
            benef_vistos = set()
            benef_vistos_F = set()
            benef_vistos_M = set()
            freq_benef = {}
            acum_pagamentos = 0
            acum_valor = 0
            
            chaves_tuple = chaves if isinstance(chaves, tuple) else (chaves,)
            base_row = dict(zip(col_agrup_sem_mes_finais, chaves_tuple))
            
            for mes_sort, mes_grupo in grupo.groupby('_sort_mes'):
                nome_mes = str(mes_grupo[col_mes_real].iloc[0])
                acum_pagamentos += len(mes_grupo)
                try:
                    acum_valor += pd.to_numeric(mes_grupo[col_valor], errors='coerce').fillna(0).sum()
                except:
                    pass
                
                for idx, row in mes_grupo.iterrows():
                    b = str(row[col_beneficiario]).strip()
                    benef_vistos.add(b)
                    freq_benef[b] = freq_benef.get(b, 0) + 1
                    
                    if col_sexo:
                        s = str(row[col_sexo]).strip().upper()
                        if s.startswith('F'): benef_vistos_F.add(b)
                        elif s.startswith('M'): benef_vistos_M.add(b)
                        
                row_dict = dict(base_row)
                row_dict[col_mes_final] = nome_mes
                
                # Copy extra string columns from relatorio_final
                mask = pd.Series(True, index=relatorio_final.index)
                for k, v in base_row.items():
                    mask &= (relatorio_final[k] == v)
                mask &= (relatorio_final[col_mes_final] == nome_mes)
                
                matched = relatorio_final[mask]
                if not matched.empty:
                    for c in colunas_string_join:
                        if c in matched.columns:
                            row_dict[c] = matched.iloc[0][c]
                            
                row_dict['Pagamentos'] = acum_pagamentos
                row_dict['Valor Pago'] = acum_valor
                row_dict['Benef. Distintos'] = len(benef_vistos)
                if col_sexo:
                    row_dict['F'] = len(benef_vistos_F)
                    row_dict['M'] = len(benef_vistos_M)
                
                contagem_freqs = {}
                for fq in freq_benef.values():
                    contagem_freqs[fq] = contagem_freqs.get(fq, 0) + 1
                    
                for i in range(1, 13):
                    row_dict[f'{i}X'] = contagem_freqs.get(i, 0)
                    row_dict[f'{i}x'] = contagem_freqs.get(i, 0)
                    
                lista_cumul.append(row_dict)
                
        if lista_cumul:
            df_cumulativo = pd.DataFrame(lista_cumul)
            
            # Check case insensitive mapping
            for c in relatorio_final.columns:
                if c not in df_cumulativo.columns:
                    mapped = False
                    for existing_col in df_cumulativo.columns:
                        if existing_col.lower() == c.lower():
                            df_cumulativo[c] = df_cumulativo[existing_col]
                            mapped = True
                            break
                    if not mapped:
                        df_cumulativo[c] = 0
            df_cumulativo = df_cumulativo[relatorio_final.columns]
            
            # Apply Window Function (cumsum) exactly as requested by user
            # This guarantees that the cumulative values are the exact running sum of the monthly table's values
            colunas_chave = col_agrup_sem_mes_finais + [col_mes_final]
            
            # Flexible lookup for 'Pagamentos' and 'Valor Pago' columns (names may vary by model)
            def _find_col(df, keywords):
                for c in df.columns:
                    c_low = str(c).lower()
                    if all(k in c_low for k in keywords):
                        return c
                return None
            
            col_pag_mensal = _find_col(relatorio_final, ['pagamento'])
            col_val_mensal = _find_col(relatorio_final, ['valor', 'pago'])
            
            # Ensure the found columns exist, otherwise fallback to direct names
            if col_pag_mensal is None:
                col_pag_mensal = 'Pagamentos' if 'Pagamentos' in relatorio_final.columns else None
            if col_val_mensal is None:
                col_val_mensal = 'Valor Pago' if 'Valor Pago' in relatorio_final.columns else None
            
            extra_cols = []
            if col_pag_mensal and col_pag_mensal in relatorio_final.columns:
                extra_cols.append(col_pag_mensal)
            if col_val_mensal and col_val_mensal in relatorio_final.columns:
                extra_cols.append(col_val_mensal)
            
            mensal_valores = relatorio_final[colunas_chave + extra_cols].copy()
            rename_mensal = {}
            if col_pag_mensal:
                rename_mensal[col_pag_mensal] = 'Pagamentos_Mensal'
            if col_val_mensal:
                rename_mensal[col_val_mensal] = 'Valor_Mensal'
            mensal_valores = mensal_valores.rename(columns=rename_mensal)
            
            # Ensure merge keys have exactly the same type (string) to prevent 'int64 vs str' merge errors
            for c in colunas_chave:
                if c in df_cumulativo.columns and c in mensal_valores.columns:
                    df_cumulativo[c] = df_cumulativo[c].astype(str)
                    mensal_valores[c] = mensal_valores[c].astype(str)
            
            df_cumulativo = pd.merge(df_cumulativo, mensal_valores, on=colunas_chave, how='left')
            
            def _get_sort_m(val):
                val_str = str(val).lower()
                meses_map = {'jan': 1, 'fev': 2, 'mar': 3, 'abr': 4, 'mai': 5, 'jun': 6, 'jul': 7, 'ago': 8, 'set': 9, 'out': 10, 'nov': 11, 'dez': 12}
                for k, v in meses_map.items():
                    if k in val_str: return v
                return 0
                
            df_cumulativo['_temp_sort_mes'] = df_cumulativo[col_mes_final].apply(_get_sort_m)
            df_cumulativo = df_cumulativo.sort_values(by=col_agrup_sem_mes_finais + ['_temp_sort_mes'])
            
            # Window Functions: SUM() OVER (PARTITION BY col_agrup_sem_mes_finais ORDER BY mes)
            if 'Pagamentos_Mensal' in df_cumulativo.columns:
                df_cumulativo['Pagamentos'] = df_cumulativo.groupby(col_agrup_sem_mes_finais)['Pagamentos_Mensal'].cumsum()
            if 'Valor_Mensal' in df_cumulativo.columns:
                df_cumulativo['Valor Pago'] = df_cumulativo.groupby(col_agrup_sem_mes_finais)['Valor_Mensal'].cumsum()
            
            drop_cols = [c for c in ['Pagamentos_Mensal', 'Valor_Mensal', '_temp_sort_mes'] if c in df_cumulativo.columns]
            df_cumulativo = df_cumulativo.drop(columns=drop_cols)
            
            for col in df_cumulativo.columns:
                if "ano" in str(col).lower():
                    df_cumulativo[col] = df_cumulativo[col].astype(str).str.replace(r'\.0$', '', regex=True).replace('nan', '')
                    
            numeric_cols = [c for c in df_cumulativo.columns if c in col_metricas and pd.api.types.is_numeric_dtype(df_cumulativo[c])]
            rename_acum = {c: f"{c}_ACUM" for c in numeric_cols}
            df_cumulativo = df_cumulativo.rename(columns=rename_acum)
        else:
            df_cumulativo = relatorio_final.copy()
    
    df_freq_renomeado = df_freq.rename(columns=rename_dict)
    meta_info = {
        'col_beneficiario': rename_dict.get(col_beneficiario, col_beneficiario),
        'col_sexo': rename_dict.get(col_sexo, col_sexo) if col_sexo else None
    }
    return df_mensal, df_cumulativo, df_freq_renomeado, meta_info

# =========================================================
# PÁGINA 1: CARREGAR & EDITAR DADOS
# =========================================================
if pagina == PAGINAS[0]:
    st.markdown("### " + "📥 1. Carregar Dados")
    st.markdown("Selecione o modelo que ditará a estrutura final dos relatórios, e de seguida faça o upload do seu ficheiro Excel com os dados em bruto.")
    
    # ---- POPUP DE ERRO DE VALIDAÇÃO ----
    if st.session_state.get("_erros_import"):
        @st.dialog("❌ Erro de Importação — Modelo Incompatível", width="large")
        def _popup_erro_import():
            modelo_err = st.session_state.get("_nome_modelo_erro", "")
            ficheiro_err = st.session_state.get("_nome_ficheiro_erro", "")
            st.markdown(f"""
**O ficheiro `{ficheiro_err}` não é compatível com o modelo selecionado: `{modelo_err}`.**

Por favor verifique se escolheu o modelo correto antes de importar.

---
**Problemas detectados:**
""")
            for err in st.session_state["_erros_import"]:
                st.markdown(f"- {err}")
            st.markdown("---")
            st.warning("⚠️ A importação foi cancelada. Corrija o modelo ou o ficheiro e tente novamente.")
            if st.button("✅ Entendido — Fechar", type="primary", use_container_width=True):
                st.session_state.pop("_erros_import", None)
                st.session_state.pop("_nome_modelo_erro", None)
                st.session_state.pop("_nome_ficheiro_erro", None)
                st.session_state.pop("_arquivo_rejeitado", None)
                st.session_state.uploader_key = st.session_state.get("uploader_key", 0) + 1
                st.rerun()
        _popup_erro_import()
    # ------------------------------------
    
    modelos_disponiveis = ["INAS", "PMA", "GIVE"] # Outros modelos podem ser adicionados aqui no futuro
    modelo_selecionado = st.selectbox("Selecione o Modelo a utilizar para os relatórios:", modelos_disponiveis)
    
    if "uploader_key" not in st.session_state:
        st.session_state.uploader_key = 0
        
    uploaded_file = st.file_uploader("Carregue o ficheiro Excel", type=["xlsx", "xls"], key=f"uploader_{st.session_state.uploader_key}")
    
    if uploaded_file is not None:
        file_model_key = f"{uploaded_file.name}_{modelo_selecionado}"
        if st.session_state.get('last_file_name') != uploaded_file.name and st.session_state.get('_arquivo_rejeitado') != file_model_key:
            with st.spinner("A analisar e processar o ficheiro..."):
                try:
                    uploaded_file.seek(0)
                    df_preview = pd.read_excel(uploaded_file, header=None, nrows=30)
                    
                    linha_cabecalho = 0
                    for i, row in df_preview.iterrows():
                        non_nulls = row.dropna()
                        if len(non_nulls) >= 3:
                            strings = [x for x in non_nulls if isinstance(x, str)]
                            if len(strings) >= len(non_nulls) * 0.5:
                                linha_cabecalho = i
                                break
                    
                    uploaded_file.seek(0)
                    df = pd.read_excel(uploaded_file, header=linha_cabecalho)
                    
                    novas_colunas = []
                    vistos = {}
                    for col in df.columns:
                        nome_limpo = str(col).replace(":", "_").strip()
                        if nome_limpo in vistos:
                            vistos[nome_limpo] += 1
                            novas_colunas.append(f"{nome_limpo}_{vistos[nome_limpo]}")
                        else:
                            vistos[nome_limpo] = 0
                            novas_colunas.append(nome_limpo)
                    df.columns = novas_colunas
                    
                    # --- VALIDAÇÃO DE MODELO ---
                    colunas_upper = [str(c).upper().strip() for c in novas_colunas]
                    
                    # Definir colunas obrigatórias por modelo
                    COLUNAS_OBRIGATORIAS = {
                        "INAS": {
                            "obrigatorias": ["PROVINCIA", "DISTRITO", "DATA"],
                            "assinatura": None,   # não exige coluna única
                            "proibidas": ["PI BARCODE", "CODIGO AGREGADO"],  # colunas que indicam outro modelo
                        },
                        "PMA": {
                            "obrigatorias": ["PI BARCODE", "PROVINCIA", "DISTRITO"],
                            "assinatura": "PI BARCODE",  # coluna-chave identificadora do PMA
                            "proibidas": [],
                        },
                        "GIVE": {
                            "obrigatorias": ["CODIGO AGREGADO", "PROVINCIA", "DISTRITO", "VALOR PAGO"],
                            "assinatura": "CODIGO AGREGADO",  # coluna exclusiva do GIVE
                            "proibidas": [],
                        }
                    }
                    
                    def _validar_modelo(modelo, colunas_upper):
                        """Retorna (ok, lista_erros)"""
                        erros = []
                        regras = COLUNAS_OBRIGATORIAS.get(modelo, {})
                        
                        # Verificar colunas obrigatórias
                        for col_req in regras.get("obrigatorias", []):
                            encontrou = any(col_req in c for c in colunas_upper)
                            if not encontrou:
                                erros.append(f"Coluna obrigatória ausente: **{col_req}**")
                        
                        # Verificar se não contém colunas que indicam outro modelo
                        for col_proib in regras.get("proibidas", []):
                            if any(col_proib in c for c in colunas_upper):
                                erros.append(f"Coluna encontrada que pertence a outro modelo: **{col_proib}**")
                        
                        # Se o modelo tem assinatura obrigatória, verificar outros modelos
                        for outro_modelo, outras_regras in COLUNAS_OBRIGATORIAS.items():
                            if outro_modelo == modelo:
                                continue
                            assinatura = outras_regras.get("assinatura")
                            if assinatura and any(assinatura in c for c in colunas_upper):
                                erros.append(f"O ficheiro parece ser do modelo **{outro_modelo}** (detectada coluna '{assinatura}'). Por favor selecione o modelo correto.")
                                break
                        
                        return len(erros) == 0, erros

                    _validacao_ok, _erros_validacao = _validar_modelo(modelo_selecionado, colunas_upper)
                    
                    if not _validacao_ok:
                        # Guardar erros no session_state para apresentar no popup
                        st.session_state["_erros_import"] = _erros_validacao
                        st.session_state["_nome_modelo_erro"] = modelo_selecionado
                        st.session_state["_nome_ficheiro_erro"] = uploaded_file.name
                        st.session_state["_arquivo_rejeitado"] = file_model_key
                        st.session_state.last_file_name = None  # bloquear re-import automático
                    # ----------------------------
                    
                    # Se validação falhou, interromper processamento
                    if not _validacao_ok:
                        st.rerun()
                    
                    # --- PRÉ-PROCESSAMENTO PMA ---
                    if modelo_selecionado == "PMA":
                        # Mapeamento dinâmico ignorando maiúsculas e espaços
                        pma_rename = {}
                        for c in df.columns:
                            cn = str(c).strip().lower()
                            if cn == 'pi barcode':
                                pma_rename[c] = 'Beneficiario'
                            elif 'delega' in cn:
                                pma_rename[c] = 'Delegação'
                            elif ('valor' in cn and 'pago' in cn) or cn == 'valores_pagos' or cn == 'valores pagos':
                                pma_rename[c] = 'Valor Pago'
                            elif ('data' in cn and 'pagamento' in cn):
                                pma_rename[c] = 'Data_Pagamento'
                            elif cn == 'n_pagamentos' or cn == 'n pagamentos' or cn == 'npagamentos' or 'pagamento' in cn and 'n' in cn:
                                pma_rename[c] = 'Pagamentos'
                                
                        df.rename(columns=pma_rename, inplace=True)
                        
                        # Injetar Datas caso vazio, conforme regra do utilizador (01/01/2026)
                        if 'Data_Pagamento' not in df.columns:
                            df['Data_Pagamento'] = '01/01/2026'
                        else:
                            df['Data_Pagamento'] = df['Data_Pagamento'].fillna('01/01/2026').replace('NaT', '01/01/2026').replace('', '01/01/2026')
                            
                        # Extrair Ano e Mês (porque o processar_relatorio agrupa por essas dimensões)
                        # Assumimos que o formador é pelo menos reconhecível por pandas, ou injetamos fixo para '01/01/2026'
                        def extrair_ano(data):
                            try:
                                return pd.to_datetime(data, dayfirst=True).year
                            except:
                                return 2026
                                
                        def extrair_mes(data):
                            meses = ['Janeiro', 'Fevereiro', 'Março', 'Abril', 'Maio', 'Junho', 'Julho', 'Agosto', 'Setembro', 'Outubro', 'Novembro', 'Dezembro']
                            try:
                                m = pd.to_datetime(data, dayfirst=True).month
                                return meses[m-1]
                            except:
                                return 'Janeiro'
                                
                        df['Ano '] = df['Data_Pagamento'].apply(extrair_ano)
                        df['Mês'] = df['Data_Pagamento'].apply(extrair_mes)
                        
                        # Forçar Implementador e Provedor
                        df['Implementador'] = 'PMA'
                        df['Provedor'] = 'Mpesa'
                    # -----------------------------
                    
                    st.session_state.df = df
                    st.session_state.df_editado = df.copy()
                    st.session_state.relatorio_final = None 
                    st.session_state.last_file_name = uploaded_file.name
                    st.session_state.modelo_selecionado = modelo_selecionado
                    
                    template_name = "modelo de tabela globalizadovf_final_xls"
                    if template_name not in st.session_state.templates:
                        template_name = list(st.session_state.templates.keys())[0]
                    template = st.session_state.templates[template_name]
                    
                    if st.session_state.get('modelo_selecionado') in ["INAS", "GIVE", "PMA"]:
                        template = template.copy()
                        template["colunas_agrupamento"] = ["Ano ", "Mês", "Província", "Distrito", "Delegação", "Fonte", "Programa", "Implementador", "Provedor  servico"]
                        template["colunas_string_join"] = []
                        template["colunas_metricas"] = ["F", "M", "Benef. Distintos", "1X", "2X", "3X", "4X", "5X", "6X", "7X", "8X", "9X", "10X", "11X", "12X", "Pagamentos", "Valor Pago"]
                    
                    df_mensal, df_cumulativo, df_bruto_mapeado, meta_info = processar_relatorio(st.session_state.df_editado.copy(), template)
                    
                    # O output do modelo PMA é 100% igual ao INAS
                    if st.session_state.get('modelo_selecionado') in ["INAS", "PMA", "GIVE"]:
                        mensal_rename_map = {
                            "Ano ": "Ano", "Mês": "Mes", "Provedor  servico": "Provedor servico",
                            "1X": "1x", "2X": "2x", "3X": "3x", "4X": "4x", "5X": "5x", "6X": "6x", 
                            "7X": "7x", "8X": "8x", "9X": "9x", "10X": "10x", "11X": "11x", "12X": "12x",
                            "ANO": "Ano", "MES": "Mes", "PROVINCIA": "Província", "DISTRITO": "Distrito",
                            "DELEGACAO": "Delegação", "FONTE": "Fonte", "PROGRAMA": "Programa",
                            "IMPLEMENTADOR": "Implementador", "PROVEDOR_SERVICO": "Provedor servico",
                            "PAGAMENTOS": "Pagamentos", "VALOR_PAGO": "Valor Pago", "BENEF_DISTINTOS": "Benef. Distintos"
                        }
                        df_mensal = df_mensal.rename(columns=mensal_rename_map)
                        df_mensal = df_mensal.loc[:, ~df_mensal.columns.duplicated()]
                        
                        cumul_rename_map = {
                            "Ano ": "Ano", "Mês": "Mes", "Provedor  servico": "Provedor servico",
                            "1X": "1x", "2X": "2x", "3X": "3x", "4X": "4x", "5X": "5x", "6X": "6x", 
                            "7X": "7x", "8X": "8x", "9X": "9x", "10X": "10x", "11X": "11x", "12X": "12x",
                            "ANO": "Ano", "MES": "Mes", "PROVINCIA": "Província", "DISTRITO": "Distrito",
                            "DELEGACAO": "Delegação", "FONTE": "Fonte", "PROGRAMA": "Programa",
                            "IMPLEMENTADOR": "Implementador", "PROVEDOR_SERVICO": "Provedor servico",
                            "F_ACUM": "F_ACUM", "M_ACUM": "M_ACUM", "Benef. Distintos_ACUM": "BENEF_DISTINTOS_ACUM",
                            "Pagamentos_ACUM": "PAGAMENTOS_ACUM", "Valor Pago_ACUM": "VALOR_PAGO_ACUM",
                            "PAGAMENTOS": "PAGAMENTOS_ACUM", "VALOR_PAGO": "VALOR_PAGO_ACUM", "BENEF_DISTINTOS": "BENEF_DISTINTOS_ACUM",
                            "Pagamentos": "PAGAMENTOS_ACUM", "Valor Pago": "VALOR_PAGO_ACUM", "Benef. Distintos": "BENEF_DISTINTOS_ACUM"
                        }
                        df_cumulativo = df_cumulativo.rename(columns=cumul_rename_map)
                        df_cumulativo = df_cumulativo.loc[:, ~df_cumulativo.columns.duplicated()]
                        
                        df_bruto_mapeado = df_bruto_mapeado.rename(columns=mensal_rename_map)
                        
                        # Para PMA a Fonte mantém-se INAS ou vazia consoante o modelo
                        if st.session_state.get('modelo_selecionado') == "PMA":
                            df_cumulativo["Fonte"] = "PMA"
                            df_mensal["Fonte"] = "PMA"
                            df_cumulativo["Implementador"] = "Mpesa"
                            df_mensal["Implementador"] = "Mpesa"
                            df_cumulativo["Provedor servico"] = ""
                            df_mensal["Provedor servico"] = ""
                        elif st.session_state.get('modelo_selecionado') == "GIVE":
                            df_cumulativo["Fonte"] = "GIVE"
                            df_mensal["Fonte"] = "GIVE"
                        else:
                            df_cumulativo["Fonte"] = "INAS"
                            df_mensal["Fonte"] = "INAS"
                        
                        # Se for PMA, forçar o implementador e provedor (apenas no caso do template não o ter feito)
                        if st.session_state.get('modelo_selecionado') == "PMA":
                            df_cumulativo["Implementador"] = "PMA"
                            df_mensal["Implementador"] = "PMA"
                            df_cumulativo["Provedor servico"] = "Mpesa"
                            df_mensal["Provedor servico"] = "Mpesa"
                        else:
                            if "Implementador" not in df_cumulativo.columns or df_cumulativo["Implementador"].astype(str).str.strip().eq("").all():
                                df_cumulativo["Implementador"] = ""
                        if "Implementador" not in df_mensal.columns or df_mensal["Implementador"].astype(str).str.strip().eq("").all():
                            df_mensal["Implementador"] = ""
                        
                        if "Programa" in df_mensal.columns and df_mensal["Programa"].astype(str).str.strip().eq("").all():
                            df_mensal["Programa"] = "PSSB"
                        elif "Programa" not in df_mensal.columns:
                            df_mensal["Programa"] = "PSSB"
                            
                        if "Programa" in df_cumulativo.columns and df_cumulativo["Programa"].astype(str).str.strip().eq("").all():
                            df_cumulativo["Programa"] = "PSSB"
                        elif "Programa" not in df_cumulativo.columns:
                            df_cumulativo["Programa"] = "PSSB"
                        
                        mensal_cols_order = ["Ano", "Mes", "Província", "Distrito", "Delegação", "Fonte", "Programa", "Implementador", "Provedor servico", "F", "M", "Benef. Distintos", "1x", "2x", "3x", "4x", "5x", "6x", "7x", "8x", "9x", "10x", "11x", "12x", "Pagamentos", "Valor Pago"]
                        for c in mensal_cols_order:
                            if c not in df_mensal.columns:
                                df_mensal[c] = 0 if c in ["F", "M", "Benef. Distintos", "Pagamentos", "Valor Pago"] or c.endswith("x") else ""
                        
                        import sys
                        if not df_mensal.columns.is_unique:
                            dups = df_mensal.columns[df_mensal.columns.duplicated()].tolist()
                            raise ValueError(f"DUPLICATE COLUMNS FOUND IN DF_MENSAL: {dups}. All columns: {df_mensal.columns.tolist()}")
                            
                        df_mensal = df_mensal[mensal_cols_order]
                        
                        cumul_cols_order = ["Ano", "Mes", "Província", "Distrito", "Delegação", "Fonte", "Programa", "Implementador", "Provedor servico", "F_ACUM", "M_ACUM", "BENEF_DISTINTOS_ACUM", "1X_ACUM", "2X_ACUM", "3X_ACUM", "4X_ACUM", "5X_ACUM", "6X_ACUM", "7X_ACUM", "8X_ACUM", "9X_ACUM", "10X_ACUM", "11X_ACUM", "12X_ACUM", "PAGAMENTOS_ACUM", "VALOR_PAGO_ACUM"]
                        for c in cumul_cols_order:
                            if c not in df_cumulativo.columns:
                                df_cumulativo[c] = 0 if "_ACUM" in c else ""
                        df_cumulativo = df_cumulativo[cumul_cols_order]
                        template["colunas_agrupamento"] = ["Ano", "Mes", "Província", "Delegação", "Distrito", "Fonte", "Programa", "Implementador", "Provedor servico"]
                        st.session_state.relatorio_final = df_mensal
                        st.session_state.relatorio_cumulativo = df_cumulativo
                    
                    # Filtros dinâmicos: apenas colunas que realmente têm dados no ficheiro
                    cols_filtradas = []
                    for c in template["colunas_agrupamento"]:
                        if c in df_mensal.columns:
                            s = df_mensal[c].astype(str).str.strip()
                            if not s.eq("").all() and not s.eq("nan").all() and not s.eq("None").all():
                                cols_filtradas.append(c)
                    st.session_state.col_agrupamento = cols_filtradas
                    st.session_state.df_bruto_mapeado = df_bruto_mapeado
                    st.session_state.meta_info = meta_info
                except Exception as e:
                    st.error(f"Erro ao ler e processar o ficheiro: {e}")
    else:
        # Ficheiro foi removido pelo utilizador, limpar tudo
        st.session_state.df_editado = None
        st.session_state.relatorio_final = None
        st.session_state.relatorio_cumulativo = None
        st.session_state.relatorio_filtrado = None
        st.session_state.relatorio_cumul_filtrado = None
        st.session_state.last_file_name = None
        st.session_state.filtros_aplicados_texto = []
        st.session_state.pop("_erros_import", None)
        st.session_state.pop("_arquivo_rejeitado", None)
        st.session_state.pop("df_bruto_mapeado", None)
        st.session_state.pop("meta_info", None)
                    
    if st.session_state.df_editado is not None:
        st.dataframe(st.session_state.df_editado, use_container_width=True, hide_index=True)
        
        st.markdown("<br>", unsafe_allow_html=True)
        col1, col2, col3 = st.columns([1, 2, 1])
        with col2:
            if st.button("Avançar para Relatórios ➡️", width="stretch", type="primary"):
                st.session_state.pagina_atual = PAGINAS[1]
                st.rerun()

# =========================================================
# PÁGINA 2: GERAR RELATÓRIO
# =========================================================
elif pagina == PAGINAS[1]:
    st.markdown("### 📑 Relatórios de Pagamentos")
    
    if st.session_state.df_editado is None:
        st.warning("⚠️ Volte à primeira página e carregue um ficheiro Excel.")
    elif st.session_state.relatorio_final is not None:
        with st.expander("🔍 Filtros de Relatório", expanded=True):
            st.markdown("<span id='filters-sticky-hook'></span>", unsafe_allow_html=True)
            rel_display = st.session_state.relatorio_final.copy()
            rel_cumul_display = st.session_state.relatorio_cumulativo.copy()
            df_bruto_mensal = st.session_state.df_bruto_mapeado.copy() if hasattr(st.session_state, 'df_bruto_mapeado') else None
            df_bruto_cumul = st.session_state.df_bruto_mapeado.copy() if hasattr(st.session_state, 'df_bruto_mapeado') else None
            meta_info = st.session_state.meta_info if hasattr(st.session_state, 'meta_info') else None
        
            filtros_aplicados = []
            if len(st.session_state.col_agrupamento) > 0:
                agrupamentos = st.session_state.col_agrupamento
                cols_per_row = 3
                
                for i in range(0, len(agrupamentos), cols_per_row):
                    row_cols = st.columns(cols_per_row)
                    for j, col in enumerate(agrupamentos[i:i+cols_per_row]):
                        with row_cols[j]:
                            is_mes = "mes" in str(col).lower() or "mês" in str(col).lower()
                            
                            if is_mes:
                                opcoes_brutas = rel_display[col].astype(str).dropna().tolist()
                                opcoes_lista = []
                                for op in opcoes_brutas:
                                    opcoes_lista.extend([m.strip() for m in op.split(',') if m.strip() and m.strip() != "N/D"])
                                try:
                                    opcoes = sorted(list(set(opcoes_lista)), key=float)
                                except:
                                    opcoes = sorted(list(set(opcoes_lista)))
                            else:
                                opcoes = sorted(list(rel_display[col].astype(str).dropna().unique()))
                                
                            selecao = st.multiselect(f"{col}", opcoes, placeholder="Seleccionar...")
                            
                            if selecao:
                                if is_mes:
                                    def _extract_mes(val_s):
                                        try:
                                            import re
                                            val_str = str(val_s).lower()
                                            nums = re.findall(r'\d+', val_str)
                                            if nums: return float(nums[0])
                                            meses_map = {'jan': 1, 'fev': 2, 'mar': 3, 'abr': 4, 'mai': 5, 'jun': 6, 'jul': 7, 'ago': 8, 'set': 9, 'out': 10, 'nov': 11, 'dez': 12}
                                            for mes_chave, num in meses_map.items():
                                                if mes_chave in val_str: return float(num)
                                            return 0.0
                                        except:
                                            return 0.0
                                            
                                    max_mes = max([_extract_mes(s) for s in selecao]) if selecao else 12
                                    
                                    mask_mensal = rel_display[col].astype(str).apply(
                                        lambda x: 0 < _extract_mes(x) <= max_mes
                                    )
                                    rel_display = rel_display[mask_mensal]
                                    
                                    if df_bruto_mensal is not None:
                                        mask_bruto_mensal = df_bruto_mensal[col].astype(str).apply(
                                            lambda x: 0 < _extract_mes(x) <= max_mes
                                        )
                                        df_bruto_mensal = df_bruto_mensal[mask_bruto_mensal]
                                    
                                    mask_cumul = rel_cumul_display[col].astype(str).apply(
                                        lambda x: 0 < _extract_mes(x) <= max_mes
                                    )
                                    rel_cumul_display = rel_cumul_display[mask_cumul]
                                    
                                    if df_bruto_cumul is not None:
                                        mask_bruto_cumul = df_bruto_cumul[col].astype(str).apply(
                                            lambda x: 0 < _extract_mes(x) <= max_mes
                                        )
                                        df_bruto_cumul = df_bruto_cumul[mask_bruto_cumul]
                                else:
                                    rel_display = rel_display[rel_display[col].astype(str).isin(selecao)]
                                    rel_cumul_display = rel_cumul_display[rel_cumul_display[col].astype(str).isin(selecao)]
                                    if df_bruto_mensal is not None:
                                        df_bruto_mensal = df_bruto_mensal[df_bruto_mensal[col].astype(str).isin(selecao)]
                                    if df_bruto_cumul is not None:
                                        df_bruto_cumul = df_bruto_cumul[df_bruto_cumul[col].astype(str).isin(selecao)]
                                    
                                filtros_aplicados.append(f"**{col}:** {', '.join(selecao)}")
        def _get_sort_mes_val(val_s):
            try:
                import re
                val_str = str(val_s).lower()
                nums = re.findall(r'\d+', val_str)
                if nums: return float(nums[0])
                meses_map = {'jan': 1, 'fev': 2, 'mar': 3, 'abr': 4, 'mai': 5, 'jun': 6, 'jul': 7, 'ago': 8, 'set': 9, 'out': 10, 'nov': 11, 'dez': 12}
                for mes_chave, num in meses_map.items():
                    if mes_chave in val_str: return float(num)
                return 0.0
            except:
                return 0.0

        col_m = next((c for c in st.session_state.col_agrupamento if "mes" in str(c).lower() or "mês" in str(c).lower()), None)
        if col_m:
            rel_display['_sort_mes'] = rel_display[col_m].apply(_get_sort_mes_val)
            rel_cumul_display['_sort_mes'] = rel_cumul_display[col_m].apply(_get_sort_mes_val)
            
            sort_cols = []
            asc_list = []
            
            # 1. Ano
            if "Ano " in st.session_state.col_agrupamento:
                sort_cols.append("Ano ")
                asc_list.append(True)
            elif "Ano" in st.session_state.col_agrupamento:
                sort_cols.append("Ano")
                asc_list.append(True)
                
            # 2. Mês (Cronológico Global)
            sort_cols.append('_sort_mes')
            asc_list.append(True)
            
            # 3. Restantes dimensões
            for c in st.session_state.col_agrupamento:
                if c != col_m and "ano" not in str(c).lower() and c in rel_display.columns:
                    sort_cols.append(c)
                    asc_list.append(True)
                    
            if sort_cols:
                rel_display = rel_display.sort_values(by=sort_cols, ascending=asc_list).drop(columns=['_sort_mes'])
                rel_cumul_display = rel_cumul_display.sort_values(by=sort_cols, ascending=asc_list).drop(columns=['_sort_mes'])

        st.session_state.relatorio_filtrado = rel_display
        st.session_state.relatorio_cumul_filtrado = rel_cumul_display
        st.session_state.filtros_aplicados_texto = filtros_aplicados
        
        with st.spinner("⏳ A calcular totais e preparar relatório..."):
            rel_display_completo, rel_display_totais = adicionar_linha_totais(
                rel_display, 
                st.session_state.col_agrupamento, 
                is_cumulativo=False, 
                df_bruto=df_bruto_mensal if 'df_bruto_mensal' in locals() else None, 
                meta=meta_info if 'meta_info' in locals() else None
            )
            
            _, rel_cumul_totais = adicionar_linha_totais(
                rel_cumul_display, 
                st.session_state.col_agrupamento, 
                is_cumulativo=True, 
                df_bruto=df_bruto_cumul if 'df_bruto_cumul' in locals() else None, 
                meta=meta_info if 'meta_info' in locals() else None
            )
            
            rel_cumul_display_completo = pd.concat([rel_cumul_display, rel_cumul_totais], ignore_index=True)
        
        def destacar_totais_isolados(row):
            is_geral = False
            for val in row.values:
                if str(val) == "TOTAL":
                    is_geral = True
                    break
            
            if is_geral:
                return ['font-weight: bold; background-color: #ffe6e6; color: black; border-top: 2px solid black'] * len(row)
            else:
                return [''] * len(row)
        
        def format_mt(val):
            if pd.isna(val) or val == "": return ""
            try:
                s = f"{float(val):,.2f}"
                s = s.replace(",", "X").replace(".", ",").replace("X", " ")
                return f"{s} MT"
            except:
                return str(val)
                
        format_dict = {col: format_mt for col in rel_display.columns if "valor" in str(col).lower() or "pago" in str(col).lower()}
        format_dict_cumul = {col: format_mt for col in rel_cumul_display.columns if "valor" in str(col).lower() or "pago" in str(col).lower()}
        
        def formatar_excel(writer, df_to_write, sheet_name, filtros):
            df_to_write.to_excel(writer, index=False, sheet_name=sheet_name)
            worksheet = writer.sheets[sheet_name]
            
            # Fixar cabeçalho
            worksheet.freeze_panes = 'A2'
            
            # Destacar linha de totais (a última linha)
            try:
                from openpyxl.styles import PatternFill, Font
                fill_totais = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
                font_totais = Font(bold=True)
                
                max_row = worksheet.max_row
                for col in range(1, worksheet.max_column + 1):
                    cell = worksheet.cell(row=max_row, column=col)
                    cell.fill = fill_totais
                    cell.font = font_totais
            except:
                pass
                
            # Adicionar filtros numa aba separada se não existir
            if filtros and "Filtros" not in writer.sheets:
                df_filtros = pd.DataFrame({"Filtros Aplicados": filtros})
                df_filtros.to_excel(writer, index=False, sheet_name="Filtros")
                try:
                    ws_filtros = writer.sheets["Filtros"]
                    ws_filtros.column_dimensions['A'].width = 100
                except:
                    pass

        def inject_df_to_template(df_to_write, template_path):
            import io
            import unicodedata
            from openpyxl import load_workbook
            
            def norm(s):
                return unicodedata.normalize('NFKD', str(s)).encode('ASCII', 'ignore').decode('utf-8').lower().strip().replace("_", " ")
                
            wb = load_workbook(template_path)
            ws = wb.active
            
            # Read template headers
            headers = []
            for cell in ws[1]:
                headers.append(cell.value)
                
            # Align dataframe to template headers
            df_aligned = pd.DataFrame()
            for h in headers:
                col_found = False
                if h is not None:
                    h_norm = norm(h)
                    for c in df_to_write.columns:
                        if norm(c) == h_norm:
                            df_aligned[h] = df_to_write[c]
                            col_found = True
                            break
                if not col_found:
                    df_aligned[h] = ""
                    
            for r_idx, row in enumerate(df_aligned.values, start=2):
                for c_idx, value in enumerate(row, start=1):
                    ws.cell(row=r_idx, column=c_idx, value=value)
                    
            buffer = io.BytesIO()
            wb.save(buffer)
            buffer.seek(0)
            return buffer

        # ---------- CRIAR RESUMO UI ----------
        # Removido agrupamento forçado por ano/província para mostrar a tabela completa na interface
        
        tab1, tab2 = st.tabs(["📊 Tabela 1 — PAGAMENTOS MENSAL", "📈 Tabela 2 — PAGAMENTOS CUMULATIVO"])
        
        with tab1:
            # Separating the UI view so the Totals row is fixed below the main scrollable table
            styled_data = rel_display.style.format(format_dict).hide(axis="index")
            styled_totais = rel_display_totais.style.apply(destacar_totais_isolados, axis=1).format(format_dict).hide(axis="index")
            
            modelo_titulo = st.session_state.get('modelo_selecionado', 'INAS')
            st.write(f"📌 **{modelo_titulo}_Relatorio mensal (Dados):**")
            st.dataframe(styled_data, use_container_width=True, hide_index=True)
            st.write("📌 **Totais (Mensal):**")
            st.dataframe(styled_totais, use_container_width=True, hide_index=True)
            
            if st.session_state.get('modelo_selecionado') == "INAS":
                template_path_mensal = r"C:\Users\administrator\Documents\xls_project\xlsx_pagamentos\INAS\Mensal.xlsx"
                with st.spinner("⏳ A gerar ficheiro Excel Mensal..."):
                    if os.path.exists(template_path_mensal):
                        buffer1 = inject_df_to_template(rel_display_completo, template_path_mensal)
                        file_name1 = "Mensal.xlsx"
                    else:
                        buffer1 = io.BytesIO()
                        with pd.ExcelWriter(buffer1, engine='openpyxl') as writer:
                            formatar_excel(writer, rel_display_completo, 'PAGAMENTOS_MENSAL', st.session_state.get('filtros_aplicados_texto', []))
                        file_name1 = "PAGAMENTOS_MENSAL.xlsx"
            else:
                with st.spinner("⏳ A gerar ficheiro Excel Mensal..."):
                    buffer1 = io.BytesIO()
                    with pd.ExcelWriter(buffer1, engine='openpyxl') as writer:
                        formatar_excel(writer, rel_display_completo, 'PAGAMENTOS_MENSAL', st.session_state.get('filtros_aplicados_texto', []))
                    file_name1 = "PAGAMENTOS_MENSAL.xlsx"
            
            st.download_button(
                label=f"📥 Descarregar Tabela 1 ({file_name1})",
                data=buffer1.getvalue(),
                file_name=file_name1,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

        with tab2:
            if st.session_state.get('modelo_selecionado') == "INAS":
                # Rebuild format dict just in case any keys were renamed
                format_dict_cumul = {col: format_dict_cumul.get(col) for col in rel_cumul_display_completo.columns if "valor" in str(col).lower() or "pago" in str(col).lower()}
                
            styled_data_cumul = rel_cumul_display.style.format(format_dict_cumul).hide(axis="index")
            styled_totais_cumul = rel_cumul_totais.style.apply(destacar_totais_isolados, axis=1).format(format_dict_cumul).hide(axis="index")
            
            modelo_titulo = st.session_state.get('modelo_selecionado', 'INAS')
            st.write(f"📌 **Acumulado_{modelo_titulo} (Dados):**")
            st.dataframe(styled_data_cumul, use_container_width=True, hide_index=True)
            st.write("📌 **Totais (Acumulado):**")
            st.dataframe(styled_totais_cumul, use_container_width=True, hide_index=True)
            
            if st.session_state.get('modelo_selecionado') == "INAS":
                template_path_cumulativo = r"C:\Users\administrator\Documents\xls_project\xlsx_pagamentos\INAS\cumulative_aggregated_payments_template.xlsx"
                with st.spinner("⏳ A gerar ficheiro Excel Acumulado..."):
                    if os.path.exists(template_path_cumulativo):
                        buffer2 = inject_df_to_template(rel_cumul_display_completo, template_path_cumulativo)
                        file_name2 = "cumulative_aggregated_payments_template.xlsx"
                    else:
                        buffer2 = io.BytesIO()
                        with pd.ExcelWriter(buffer2, engine='openpyxl') as writer:
                            formatar_excel(writer, rel_cumul_display_completo, 'PAGAMENTOS_CUMULATIVO', st.session_state.get('filtros_aplicados_texto', []))
                        file_name2 = "PAGAMENTOS_CUMULATIVO.xlsx"
            else:
                with st.spinner("⏳ A gerar ficheiro Excel Acumulado..."):
                    buffer2 = io.BytesIO()
                    with pd.ExcelWriter(buffer2, engine='openpyxl') as writer:
                        formatar_excel(writer, rel_cumul_display_completo, 'PAGAMENTOS_CUMULATIVO', st.session_state.get('filtros_aplicados_texto', []))
                    file_name2 = "PAGAMENTOS_CUMULATIVO.xlsx"
            
            st.download_button(
                label=f"📥 Descarregar Tabela 2 ({file_name2})",
                data=buffer2.getvalue(),
                file_name=file_name2,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

# =========================================================
# PÁGINA 3: DASHBOARD VISUAL
# =========================================================
elif pagina == PAGINAS[2]:
    st.markdown("### " + "📈 3. Dashboard Visual")
    
    import plotly.express as px
    import plotly.graph_objects as go
    
    if st.session_state.relatorio_cumulativo is None:
        st.warning("⚠️ Primeiro, vá à secção 'Gerar Relatório' e processe a tabela.")
    else:
        with st.spinner("⏳ A carregar dashboard..."):
            # 1. Obter Tabela Cumulativa (com ou sem filtros da Página 2)
            if "relatorio_cumul_filtrado" in st.session_state and st.session_state.relatorio_cumul_filtrado is not None:
                rel_viz = st.session_state.relatorio_cumul_filtrado.copy()
            else:
                rel_viz = st.session_state.relatorio_cumulativo.copy()
        
        filtros_txt = st.session_state.get('filtros_aplicados_texto', [])
        st.write("*(Nota: Os gráficos respondem aos filtros aplicados na página '2. Gerar Relatório'.)*")
        if filtros_txt:
            st.info("🎯 **Filtros Ativos:** " + " | ".join(filtros_txt))
        else:
            st.info("🎯 **Filtros Ativos:** Nenhum (A mostrar todos os dados acumulados)")
            
        st.divider()
        
        # 2. Calcular KPIs (Para dados cumulativos, o "Total" correto é o máximo de cada partição,
        # que é o que a função adicionar_linha_totais já faz por nós)
        df_bruto_atual = st.session_state.df_bruto_mapeado if hasattr(st.session_state, 'df_bruto_mapeado') else None
        meta_info = st.session_state.meta_info if hasattr(st.session_state, 'meta_info') else None
        
        # Garantir que temos col_agrupamento
        col_agrupamento = st.session_state.get('col_agrupamento', [])
        
        _, df_totais = adicionar_linha_totais(
            rel_viz, 
            col_agrupamento, 
            is_cumulativo=True, 
            df_bruto=df_bruto_atual, 
            meta=meta_info
        )
        
        # 3. Extrair Valores para os KPIs
        def extract_kpi(col_name):
            if col_name in df_totais.columns:
                val = df_totais[col_name].iloc[0]
                return pd.to_numeric(val, errors='coerce') if val != '' else 0
            return 0

        total_valor = extract_kpi('VALOR_PAGO_ACUM')
        total_benef = extract_kpi('BENEF_DISTINTOS_ACUM')
        total_pags = extract_kpi('PAGAMENTOS_ACUM')
        total_f = extract_kpi('F_ACUM')
        total_m = extract_kpi('M_ACUM')
        
        # Encontrar total outros generos (se existirem)
        total_outros = 0
        for col in df_totais.columns:
            if col.startswith("Sexo_") or col == "Sem_Gênero":
                total_outros += extract_kpi(col)
        
        # Formatador de MT
        def formata_mt_kpi(val):
            if pd.isna(val): return "0 MT"
            s = f"{float(val):,.2f}"
            return s.replace(",", "X").replace(".", ",").replace("X", " ") + " MT"
            
        # 4. Renderizar Cartões (KPIs)
        kpi1, kpi2, kpi3 = st.columns(3)
        kpi1.metric("💰 Valor Distribuído Acumulado", formata_mt_kpi(total_valor))
        kpi2.metric("👥 Beneficiários Únicos", f"{float(total_benef):,.0f}")
        kpi3.metric("💳 Pagamentos Efetuados", f"{float(total_pags):,.0f}")
        
        st.write("")
        
        kpi_m, kpi_f, kpi_o = st.columns(3)
        kpi_m.metric("👨 Homens (M)", f"{float(total_m):,.0f}")
        kpi_f.metric("👩 Mulheres (F)", f"{float(total_f):,.0f}")
        kpi_o.metric("👤 Gênero Não Informado", f"{float(total_outros):,.0f}")
        
        st.divider()
        
        # 5. Gráficos Plotly
        
        # Descobrir a coluna de Mês
        col_m = next((c for c in col_agrupamento if "mes" in str(c).lower() or "mês" in str(c).lower()), None)
        
        if col_m and not rel_viz.empty:
            # Ordenar temporariamente pelos meses para o gráfico de linha/área
            def _get_sort_m(val):
                val_str = str(val).lower()
                meses_map = {'jan': 1, 'fev': 2, 'mar': 3, 'abr': 4, 'mai': 5, 'jun': 6, 'jul': 7, 'ago': 8, 'set': 9, 'out': 10, 'nov': 11, 'dez': 12}
                for k, v in meses_map.items():
                    if k in val_str: return v
                return 0
                
            rel_viz['_sort_m'] = rel_viz[col_m].apply(_get_sort_m)
            graf_tempo = rel_viz.groupby([col_m, '_sort_m'])[['VALOR_PAGO_ACUM', 'PAGAMENTOS_ACUM']].sum().reset_index()
            graf_tempo = graf_tempo.sort_values('_sort_m')
            
            if not graf_tempo.empty:
                st.markdown("#### 📈 Evolução do Valor Pago (Acumulado)")
                fig_area = px.area(
                    graf_tempo, 
                    x=col_m, 
                    y='VALOR_PAGO_ACUM',
                    labels={'VALOR_PAGO_ACUM': 'Valor Distribuído (MT)', col_m: 'Mês'},
                    color_discrete_sequence=["#0083B8"]
                )
                fig_area.update_layout(margin=dict(l=0, r=0, t=30, b=0), xaxis_title=None, height=280)
                st.plotly_chart(fig_area, use_container_width=True)
                
        st.write("")
        
        # Gráficos Secundários
        c1, c2 = st.columns(2)
        
        with c1:
            st.markdown("#### 🚻 Proporção por Gênero")
            labels_sexo = []
            valores_sexo = []
            
            if total_f > 0:
                labels_sexo.append('Mulheres (F)')
                valores_sexo.append(total_f)
            if total_m > 0:
                labels_sexo.append('Homens (M)')
                valores_sexo.append(total_m)
            if total_outros > 0:
                labels_sexo.append('Outros')
                valores_sexo.append(total_outros)
                
            if sum(valores_sexo) > 0:
                fig_pie = px.pie(
                    names=labels_sexo, 
                    values=valores_sexo, 
                    hole=0.4,
                    color_discrete_sequence=["#FF4B4B", "#0083B8", "#888888"]
                )
                fig_pie.update_layout(margin=dict(l=0, r=0, t=30, b=0), height=250)
                st.plotly_chart(fig_pie, use_container_width=True)
            else:
                st.info("Sem dados de gênero para exibir.")
                
        with c2:
            freq_cols = [c for c in ['1X_ACUM', '2X_ACUM', '3X_ACUM', '4X_ACUM', '5X_ACUM', '6X_ACUM', '7X_ACUM', '8X_ACUM', '9X_ACUM', '10X_ACUM', '11X_ACUM', '12X_ACUM'] if c in df_totais.columns]
            if freq_cols:
                st.markdown("#### 🔄 Frequência de Pagamentos (Vezes)")
                freq_vals = [extract_kpi(c) for c in freq_cols]
                
                # Filtrar os que têm > 0
                freq_data = pd.DataFrame({'Vezes': [c.replace('_ACUM', '') for c in freq_cols], 'Total': freq_vals})
                freq_data = freq_data[freq_data['Total'] > 0]
                
                if not freq_data.empty:
                    fig_bar = px.bar(
                        freq_data, 
                        y='Vezes', 
                        x='Total', 
                        orientation='h',
                        text='Total',
                        color_discrete_sequence=["#28a745"]
                    )
                    fig_bar.update_layout(margin=dict(l=0, r=0, t=30, b=0), yaxis={'categoryorder':'total ascending'}, xaxis_title=None, yaxis_title=None, height=250)
                    st.plotly_chart(fig_bar, use_container_width=True)
                else:
                    st.info("Sem dados de frequência para exibir.")
                    
        st.write("")
        st.divider()
        
        # Treemap por Distrito / Província
        if col_agrupamento:
            eixo_x = col_agrupamento[-1]
            for col in col_agrupamento:
                if "distrito" in col.lower() or "deleg" in col.lower() or "prov" in col.lower():
                    eixo_x = col
                    break
                    
            st.markdown(f"#### 📍 Distribuição Geográfica de Pagamentos Acumulados ({eixo_x})")
            if not rel_viz.empty and 'VALOR_PAGO_ACUM' in rel_viz.columns:
                # Group by to find max value per district
                graf_geo = rel_viz.groupby(eixo_x)['VALOR_PAGO_ACUM'].max().reset_index()
                graf_geo = graf_geo[graf_geo['VALOR_PAGO_ACUM'] > 0]
                
                if not graf_geo.empty:
                    fig_tree = px.treemap(
                        graf_geo, 
                        path=[eixo_x], 
                        values='VALOR_PAGO_ACUM',
                        color='VALOR_PAGO_ACUM',
                        color_continuous_scale='Blues'
                    )
                    fig_tree.update_layout(margin=dict(l=0, r=0, t=30, b=0), height=300)
                    st.plotly_chart(fig_tree, use_container_width=True)
                else:
                    st.info("Sem dados geográficos para exibir.")

# =========================================================
# BOTÕES DE NAVEGAÇÃO DE PÁGINA (Fixos em baixo)
# =========================================================# =========================================================
# BOTÕES DE NAVEGAÇÃO DE PÁGINA (Fixos em baixo)
# =========================================================


col1, col2, col3, col4, col5 = st.columns([1, 1, 1, 1, 1])

with col1:
    if st.session_state.pagina_atual != PAGINAS[0]:
        st.button("⬅️ Voltar", on_click=ir_anterior, width="stretch")

with col5:
    if st.session_state.pagina_atual != PAGINAS[-1]:
        pode_avancar = True
        if st.session_state.pagina_atual == PAGINAS[0] and st.session_state.df_editado is None:
            pode_avancar = False
        elif st.session_state.pagina_atual == PAGINAS[1] and st.session_state.relatorio_final is None:
            pode_avancar = False
            
        st.button("Avançar ➡️", on_click=ir_proximo, width="stretch", type="primary", disabled=not pode_avancar)


st.markdown("<div id='nav-buttons-hook'></div>", unsafe_allow_html=True)
